#!/usr/bin/python
'''
A collection of useful tools and functions for the manipulation of SBtab tables and SBtab documents.
'''
import re
import string
import libsbml
import numpy
import scipy
import scipy.linalg
import scipy.optimize
import random
import copy
import math
import os

try: from . import SBtab
except: import SBtab


def table_type(sbtab):
    '''
    determines table_type of SBtab file
    '''
    for row in sbtab.split('\n'):
        if row.startswith('!!'):
            try:
                row = row.replace('"', "'")
                tabletype = re.search("TableType='([^']*)'", row).group(1)
                return tabletype
            except: pass
    return False


def count_tabs(sbtab_string):
    '''
    Counts how many SBtabs are in in a given string.

    Parameters
    ----------
    sbtab_string: str
        SBtab table or tables in string representation.
        
    Returns: int
        Amount of SBtab tables in given string.
    '''
    counter = 0
    for row in sbtab_string.split('\n'):
        if row.startswith('!!SBtab') or row.startswith('!!ObjTables'):
            counter += 1
    return counter


def validate_file_extension(file_name, file_type):
    '''
    Returns Boolean flag to evaluate if the file has the correct extension:
    sbml => xml
    sbtab => tsv, csv, xlsx.

    Parameters
    ----------
    file_name: str
        Name of the file.
    file_type: str
        Type of the file ('sbtab' or 'sbml').

    Returns: Bool
        Boolean flag indicating if the given file extension corresponds to the given file type.
    '''
    # check extension for sbml file
    if file_type == 'sbml' and file_name[-3:] == 'xml': return True
    elif file_type == 'sbml': return False
    else: pass

    # check extension for sbtab file
    if file_type == 'sbtab' and file_name[-3:] == 'tsv': return True
    elif file_type == 'sbtab' and file_name[-3:] == 'csv': return True
    elif file_type == 'sbtab' and file_name[-4:] == 'xlsx': return True
    elif file_type == 'sbtab': return False
    else: pass

    # if something is completely off, return False
    return False


def check_delimiter(sbtab_file):
    '''
    Determines the delimiter of the SBtab table

    Parameters
    ----------
    sbtab_file: str
        SBtab table in string representation.

    Returns: str
        Delimiter of the SBtab table ('\t', ',', or ';')
    '''
    sep = False

    try:
        for row in sbtab_file.split('\n'):
            if row.startswith('!!'): continue
            if row.startswith('!'):
                s = re.search('(.)(!)', row[1:])
                # if there is only 1 column, we have to define a default separator
                # let's use a tab.
                try: sep = s.group(1)
                except: sep = '\t'
    except: pass

    return sep


def valid_prior(sbtab_prior):
    '''
    if the given SBtab file is a prior for parameter balancing, it needs to be
    checked thorougly for the validity of several features
    '''
    validity = []

    # check table type
    if sbtab_prior.table_type != 'QuantityInfo':
        validity.append('Error: The TableType of the prior file is not '\
                        'correct: %s. '\
                        'It should be QuantityInfo' % sbtab_prior.table_type)

    # check for required columns
    required_columns = ['!QuantityType', '!Unit', '!MathematicalType',
                        '!PriorMedian', '!PriorStd', '!PriorGeometricStd',
                        '!DataStd', '!Dependence',
                        '!UseAsPriorInformation', '!MatrixInfo']
    for column in required_columns:
        if column not in sbtab_prior.columns_dict:
            validity.append('Error: The crucial column %s is missing in'\
                            ' the prior file.' % column)

    # check for required row entries
    required_rows = ['standard chemical potential',
                     'catalytic rate constant geometric mean',
                     'concentration', 'concentration of enzyme',
                     'Michaelis constant', 'inhibitory constant',
                     'activation constant', 'chemical potential',
                     'product catalytic rate constant',
                     'substrate catalytic rate constant',
                     'equilibrium constant', 'forward maximal velocity',
                     'reverse maximal velocity', 'reaction affinity']
    for row in sbtab_prior.value_rows:
        try: required_rows.remove(row[sbtab_prior.columns_dict['!QuantityType']])
        except: pass

    for row in required_rows:
        validity.append('Error: The prior file is missing an entry for th'\
                        'crucial value %s.' % row)

    return validity


def extract_pseudos_priors(sbtab_prior):
    '''
    extracts the priors and pseudos of a given SBtab prior table
    '''
    pseudo_list = ['chemical potential', 'product catalytic rate constant',
                   'substrate catalytic rate constant',
                   'equilibrium constant', 'forward maximal velocity',
                   'reverse maximal velocity', 'reaction affinity']
    pmin = {}
    pmax = {}
    pseudos = {}
    priors = {}
    
    for row in sbtab_prior.value_rows:
        pmin[row[sbtab_prior.columns_dict['!QuantityType']]] = float(row[sbtab_prior.columns_dict['!LowerBound']])
        pmax[row[sbtab_prior.columns_dict['!QuantityType']]] = float(row[sbtab_prior.columns_dict['!UpperBound']])
        if row[sbtab_prior.columns_dict['!MathematicalType']] == 'Additive':
            std = row[sbtab_prior.columns_dict['!PriorStd']]
        else:
            std = row[sbtab_prior.columns_dict['!PriorGeometricStd']]
        median = row[sbtab_prior.columns_dict['!PriorMedian']]

        if row[sbtab_prior.columns_dict['!QuantityType']] in pseudo_list:
            pseudos[row[sbtab_prior.columns_dict['!QuantityType']]] = [float(median),
                                                                       float(std)]
        else:
            priors[row[sbtab_prior.columns_dict['!QuantityType']]] = [float(median),
                                                                      float(std)]
       
    return pseudos, priors, pmin, pmax


def readout_config(sbtab_options):
    '''
    reads out the content of an optional config file and returns a parameter
    dictionary with many options for the balancing process
    '''
    parameter_dict = {'config': True}
    log = []
    allowed_options = ['use_pseudo_values', 'ph', 'temperature',
                       'overwrite_kinetics', 'cell_volume', 'parametrisation',
                       'enzyme_prefactor', 'default_inhibition',
                       'default_activation', 'model_name', 'boundary_values',
                       'samples', 'size_limit']


    if '!ID' not in sbtab_options.columns_dict:
        log.append('Error: The crucial option (ID) column is missing from the'\
                   'options file')
    if '!Value' not in sbtab_options.columns_dict:
            log.append('Error: The crucial value column is missing from the'\
                       'options file')

    for row in sbtab_options.value_rows:
        if row[sbtab_options.columns_dict['!ID']] not in allowed_options:
            log.append('There was an irregular option in the options file:'\
                       '%s' % row[sbtab_options.columns_dict['!ID']])
        else:
            if row[sbtab_options.columns_dict['!Value']] == '':
                log.append('There is no value set for option:'\
                           '%s' % row[sbtab_options.columns_dict['!ID']])
    
        parameter_dict[row[sbtab_options.columns_dict['!ID']]] = row[sbtab_options.columns_dict['!Value']]

    return parameter_dict, log


def size_warning(sbml_file):
    '''
    check whether a given sbml file has more than 250 reactions and 
    then yield a warning message for the online interface
    '''
    reader = libsbml.SBMLReader()
    sbml = reader.readSBMLFromString(sbml_file)
    sbml_model = sbml.getModel()
    if sbml_model.getNumReactions() > 250:
        return ('Warning: The model has more than 250 reactions, which may'\
                ' slow down the computation time significantly. Proceed'\
                ' with care.')
    else: return False    


def get_modifiers(reaction):
    '''
    get iterator for the reaction modifiers
    @type reaction: libsbml.Reaction
    @rtype: Iterator
    '''
    for s in reaction.getListOfModifiers():
        yield s


def get_participants(reaction):
    """
    get iterator for the reaction participants (reactants + products)
    @type reaction: libsbml.Reaction
    @rtype: Iterator
    """
    for s in reaction.getListOfReactants():
        yield s
    for s in reaction.getListOfProducts():
        yield s


def is_enzyme(species):
    return species.getSBOTerm() == 14 or species.getId().startswith('enzyme') \
        or species.getSBOTerm() == 460


def get_enzyme_for_reaction(reaction, create=False):
    is_enzyme = lambda s: s.getSBOTerm() == 14 \
                or s.getId().startswith('enzyme') or s.getSBOTerm() == 460
    for m in reaction.getListOfModifiers():
        s = reaction.getModel().getSpecies(m.getSpecies())
        if is_enzyme(s):
            return s
    if create:
        e = reaction.getModel().createSpecies()
        e.setId('enzyme_' + reaction.getId())
        e.setName('enzyme_' + reaction.getId())
        try:
            comp = reaction.getModel().getSpecies(reaction.getReactant(0).getSpecies()).getCompartment()
        except:
            comp = reaction.getModel().getSpecies(reaction.getProduct(0).getSpecies()).getCompartment()
        if not comp:
            comp = reaction.getModel().getCompartment(0).getId()
        e.setCompartment(comp)
        e.setSBOTerm(460)
        mod = reaction.createModifier()
        mod.setSpecies(e.getId())
        return e
    raise Exception('No enzyme was found for reaction %s' % reaction.getId())


def fmin_gen(f, x0, population_size=100, survivors=20, generations=20000,
             bounds=None, variable_is_logarithmic=None, intruders=0,
             use_pp=True, convenience_class=None, disp=1):
    import struct

    f = open('medians.txt', 'r')
    medians_no = []
    content = f.read()
    for elem in content.split(','):
        medians_no.append(float(elem))
    f.close()

    g = open('cpost.txt', 'r')
    content2 = g.read()
    C_post_no = []
    for line in content2.split('\n'):
        if line != '':
            linecontent = line.split(',')
            single_line = []
            for elem in linecontent:
                single_line.append(float(elem))
            C_post_no.append(single_line)

    medians = numpy.array(medians_no)
    C_post = numpy.array(C_post_no)

    def local_optimize(indiv, convenience_class=None):
        better_indiv = indiv
        if convenience_class:
            better_indiv = scipy.optimize.fmin_bfgs(convenience_class.f,
                                                    better_indiv, disp=0,
                                                    maxiter=20)
            fval = convenience_class.f_opt(better_indiv, medians, C_post)
        else:
            fval = f_opt(better_indiv, medians, C_post)
        return [better_indiv, fval]

    def float_to_bits(value):
        if type(value) == float:
            return (str(struct.unpack('Q',
                                      struct.pack('d',
                                                  value))[0])).rjust(20, "0")
        else:
            return ",".join([float_to_bits(x) for x in value.tolist()])

    def bits_to_float(bits):
        if "," in bits:
            return scipy.array([bits_to_float(x) for x in bits.split(",")])
        else:
            return struct.unpack('d', struct.pack('Q', int(bits)))[0]

    def new_individual():
        x = []
        for i in range(indiv_size):
            if variable_is_logarithmic[i]:
                logmin = scipy.log(bounds[i][0])
                logmax = scipy.log(bounds[i][1])
                x.append(scipy.exp(scipy.rand() * (logmax - logmin) + logmin))
            else:
                x.append(scipy.rand() * (bounds[i][1] - bounds[i][0]) +
                         bounds[i][0])
        return scipy.array(x)

    def bool_mate(mother, father):
        ms = float_to_bits(mother)
        fs = float_to_bits(father)
        l = random.sample(range(len(ms)), 3)
        l.sort()
        [i1, i2, i3] = l
        cs = ms[:i1] + fs[i1:i2] + ms[i2:i3] + fs[i3:]
        child = bits_to_float(cs)
        if child.size != mother.size or None in child or scipy.inf in child \
           or -scipy.inf in child: raise ValueError()
        return child

    def mate(mflist):
        return mflist[0] + mflist[1] - mflist[2]

    def mutate(indiv):
        if not False:
            bi = float_to_bits(indiv)
            number = int(math.ceil(float(len(bi)) / 100.))
            change_indices = random.sample([x for x in range(3, len(bi))],
                                           number)
            for ci in change_indices:
                new = bi[:ci] + str(random.choice(range(10))) + bi[(ci + 1):]
                try:
                    bits_to_float(new)
                    bi = new
                except struct.error:
                    # dont accept change
                    pass
            return scipy.absolute(bits_to_float(bi))
        else:
            return scipy.exp(scipy.log(indiv) + wolf["mutation_factor"] *
                             scipy.array([random.normalvariate(0, 1) for x in range(indiv_size)]))

    def bound(vector):
        global correct_vector

        if len(vector) == indiv_size:
            correct_vector = vector

        if len(vector) != indiv_size:
            print('im doing it right now!')
            vector = correct_vector

        for i in range(indiv_size):
            if vector[i] < bounds[i][0]:
                vector[i] = bounds[i][0]
                correct_vector = vector
            elif vector[i] > bounds[i][1]:
                vector[i] = bounds[i][1]
                correct_vector = vector
        return vector

    if bounds is None:
        bounds = [[1e-4, 1e4]] * len(x0)
    if len(bounds) != len(x0):
        raise Exception('Length of x0 and length of bounds do not fit!')
    if variable_is_logarithmic is None:
        variable_is_logarithmic = [[True]] * len(x0)
    if len(variable_is_logarithmic) != len(x0):
        raise Exception('Length of variable_is_logarithmic and x0 do not fit!')

    if use_pp:
        import pp
        # get pp servers
        servers = []
        fi = file("servers", "r")
        for l in fi:
            if l.startswith("#"): continue
            servers.append(l.split(",")[1])
        fi.close()
        # read secret file
        try:
            fi = open("secret", "r")
            secret = fi.readlines()[0]
            fi.close()
        except:
            import sys
            print('''Please create a file called \"secret\" with one long
            string in the first line.''')
            sys.exit(1)
        if disp == 1:
            print("starting servers", servers)
        job_server = pp.Server(ppservers=tuple(servers), secret=secret)
        print(job_server.get_active_nodes())

    population = [x0]
    indiv_size = x0.size
    quality = []
    for i in range(population_size - len(population)):
        population.append(new_individual())
    best_indiv = copy.deepcopy(x0)

    try:
        for i in range(generations):
            # rate individuals
            local_optimization_results = []
            pre_computed_qualities = len(quality)

            for j in range(pre_computed_qualities, population_size):
                if use_pp:
                    if convenience_class:
                        local_optimization_results.append(job_server.submit(local_optimize,
                                                                            (population[j],
                                                                             convenience_class),
                                                                            modules=("scipy",
                                                                                     "scipy.optimize",
                                                                                     "scipy.linalg",
                                                                                     "bounded"),
                                                                            globals=globals()))
                    else:
                        local_optimization_results.append(job_server.submit(local_optimize,
                                                                            (population[j],),
                                                                            modules=("scipy",
                                                                                     "scipy.optimize",
                                                                                     "scipy.linalg",
                                                                                     "bounded"),
                                                                            globals=globals()))
                else:
                    local_optimization_results.append(local_optimize(population[j]))

            if use_pp:
                for j in range(len(local_optimization_results)):
                    local_optimization_results[j] = local_optimization_results[j]()
            for j in range(len(local_optimization_results)):
                [better_indiv, fval] = local_optimization_results[j]
                population[pre_computed_qualities + j] = better_indiv
                quality.append(fval)

            # replace None results
            for j in range(len(population)):
                while quality[j] is None or scipy.isnan(quality[j]):
                    population[j] = new_individual()
                    quality[j] = f_opt(population[j], medians, C_post)

            # sort
            sorted_quality = list(zip(quality, population))
            sorted_quality.sort(key=lambda x: x[0])
            new_population = []
            new_quality = []
            for j in range(survivors):
                f_val, indiv = sorted_quality[j]
                new_population.append(indiv)
                new_quality.append(f_val)
            population = new_population
            quality = new_quality

            # intrude
            for j in range(intruders):
                population.append(new_individual())

            # mate
            current_size = len(population)
            for j in range(current_size, population_size):
                population.append(bound(mutate(mate(random.sample(population[:current_size],
                                                                  3)))))
            best_indiv = copy.deepcopy(population[0])
            if disp == 1:
                print("generation", str(i + 1).rjust(7), "     f =",
                      quality[0])
                print(best_indiv)
    except KeyboardInterrupt:
        if disp == 1:
            print("Stopping computation")

    if use_pp:
        job_server.print_stats()
    if disp == 1:
        print("generation goodbye      f =", quality[0])

    return best_indiv


def fmin_differential_evolution(f, x0, population_size=100, generations=20000,
                                bounds=None, variable_is_logarithmic=None,
                                crossover_factor=0.2, disp=0):
    import struct

    p = open('medians.txt', 'r')
    medians_no = []
    content = p.read()
    for elem in content.split(','):
        medians_no.append(float(elem))
    p.close()

    g = open('cpost.txt', 'r')
    content2 = g.read()
    C_post_no = []
    for line in content2.split('\n'):
        if line != '':
            linecontent = line.split(',')
            single_line = []
            for elem in linecontent:
                single_line.append(float(elem))
            C_post_no.append(single_line)

    medians = numpy.array(medians_no)
    C_post = numpy.array(C_post_no)

    def float_to_bits(value):
        if type(value) == float:
            return (str(struct.unpack('Q', struct.pack('d',
                                                       value))[0])).rjust(20,
                                                                          "0")
        else:
            return ",".join([float_to_bits(x) for x in value.tolist()])

    def bits_to_float(bits):
        if "," not in bits:
            return struct.unpack('d', struct.pack('Q', long(bits)))[0]
        else:
            return scipy.array([bits_to_float(x) for x in bits.split(",")])

    def new_individual():
        x = []
        for i in range(indiv_size):
            if variable_is_logarithmic[i]:
                logmin = scipy.log(bounds[i][0])
                logmax = scipy.log(bounds[i][1])
                x.append(scipy.exp(scipy.rand() * (logmax - logmin) + logmin))
            else:
                x.append(scipy.rand() * (bounds[i][1] - bounds[i][0]) +
                         bounds[i][0])
        return scipy.array(x)

    def crossover(orig, crossing_vector):
        for i in range(len(orig.tolist())):
            if random.random() < crossover_factor:
                orig[i] = crossing_vector[i]

    def bound(vector):
        for i in range(indiv_size):
            if vector[i] < bounds[i][0]:
                vector[i] = bounds[i][0]
            elif vector[i] > bounds[i][1]:
                vector[i] = bounds[i][1]

    if bounds is None:
        bounds = [[1e-4, 1e4]] * len(x0)
    if len(bounds) != len(x0):
        raise Exception('Length of x0 and length of bounds do not fit!')
    if variable_is_logarithmic is None:
        variable_is_logarithmic = [[True]] * len(x0)
    if len(variable_is_logarithmic) != len(x0):
        raise Exception('Length of variable_is_logarithmic and x0 do not fit!')

    population = [x0]
    indiv_size = x0.size
    quality = [f(x0)]

    for i in range(population_size - len(population)):
        population.append(new_individual())
        quality.append(f(population[-1]))
    best_indiv = copy.deepcopy(x0)

    try:
        for i in range(generations):
            for targetindex in range(len(population)):
                [v1, v2, v3] = random.sample(population, 3)
                trial_vector = v1 + v2 - v3
                target_vector = population[targetindex]
                crossover(trial_vector, target_vector)
                bound(trial_vector)
                trial_quality = f(trial_vector)
                if trial_quality < quality[targetindex]:
                    population[targetindex] = trial_vector
                    quality[targetindex] = trial_quality

            # replace None results
            for j in range(len(population)):
                while quality[j] is None:
                    population[j] = new_individual()
                    quality[j] = f(population[j])

            best_index = quality.index(min(quality))
            if disp == 1:
                print("generation", str(i + 1).rjust(7), "     f =",
                      quality[best_index])
                print(population[best_index])
            best_indiv = copy.deepcopy(population[best_index])

    except KeyboardInterrupt:
        if disp == 1:
            print("Stopping computation")

    return best_indiv


def split_sbtabs(sbtab_strings):
    '''
    Cuts one SBtab string in single SBtab strings if necessary.

    Parameters
    ----------
    sbtab_strings: str
        SBtab table or tables in string representation.

    Returns: list
        List of SBtab tables in string representation.
    '''
    sbtabs = []
    sbtab_string = ''
    counter = 1
    
    for row in sbtab_strings.split('\n'):
        if row.startswith('!!!') or row.startswith('"!!!'): continue
        if row.startswith('!!'):
            if sbtab_string == '':
                sbtab_string = row + '\n'
                continue
            else:
                try:
                    if sbtab_string.startswith('!!SBtab') or sbtab_string.startswith('!!ObjTables'):
                        sbtabs.append(sbtab_string)
                        counter += 1
                    sbtab_string = row + '\n'
                except:
                    print('Warning: Could not write SBtab %s' % counter)
                    counter += 1
        else:
            sbtab_string += row + '\n'

    if sbtab_string.startswith('!!SBtab') or sbtab_string.startswith('!!ObjTables'):
        sbtabs.append(sbtab_string)
                    
    return sbtabs


def sbtab_to_html(sbtab, filename=None, mode='sbtab_online', template = [], put_links = True, title_string='', show_header_row = True, show_table_name = False, show_table_text = False, definitions_file=''):
    '''
    Generates html view out of SBtab table or SBtab document object.

    Parameters
    ----------
    sbtab: SBtab.SBtabTable | SBtab.SBtabDocument
        Either SBtab table object or SBtab document object.
    filename: str
        File name of the SBtab table.
    mode: str
        Defines the type of HTML to be generated ('sbtab_online' for the SBtab online
        interface or 'standalone' for a sole HTML page without online binding).

    Returns: str
        SBtab object as HTML string.
    '''
    def _is_float(value):
        '''
        checks if an element is a float in string format
        '''
        try:
            float(value)
            return True
        except: return False
        
    def _build_main(sbtab, sbtab_def):
        '''
        builds main body of HTML, which needs to be repeated
        for SBtab Documents
        '''
        no_link = ['(',')','+','-','<=>','or','and','FbcOr','FbcAnd']
        # get column descriptions for this table type and possible shortname links
        try: (col2description,col2link) = find_descriptions(sbtab_def, sbtab.table_type)
        except:
            col2description = False
            col2link = False
        
        # start main
        html = '<table class="table-striped">'

        # table name
        if show_table_name:
            html += '<center><h2>%s</h2></center>' % (sbtab.get_attribute('TableName'))

        if show_table_text:
            if len(sbtab.get_attribute('Text')):
                html += '<center><p>%s</p></center>' % (sbtab.get_attribute('Text'))

        # header row
        #html += '<thead><tr><th colspan="%s">%s</th></tr></thead>' % (len(sbtab.columns), sbtab.header_row)
        if show_header_row:
            html += '<h4>%s</h4>' % (sbtab.header_row)

        # columns
        html += '<thead>'
        html += '<tr style="line-height:2;">'
        for col in sbtab.columns:
            try: title = col2description[col[1:]]
            except: title = ''
            html += '<th title="%s">%s</th>' % (title, col)
        html += '</tr>'
        html += '</thead>'
        html += '<tbody>'

        # value rows
        for row in sbtab.value_rows:
            # set anchor for internal jump links
            try: html += '<tr id="%s" style="line-height:1.5;">' % row[sbtab.columns_dict['!ID']]
            except: html += '<tr style="line-height:1.5;">'
            for i,col in enumerate(row):
                # try and set internal jump links via shortnames
                try:
                    col2link[sbtab.columns[i]]
                    if col2link[sbtab.columns[i]] == 'True' and col != '' and col != False:
                        try:
                            html += '<td>'
                            split_column = col.split(' ')
                            for element in split_column:
                                if element not in no_link and not _is_float(element) and put_links:
                                    #html += '<a href="#%s">%s</a> ' % (element, element)    #internal links
                                    html += element
                                else:
                                    html += element + ' '
                            html += '</td>'
                        except: html += '<td>%s</td>' % (col)
                    else:
                        html += '<td>%s</td>' % (col)
                except:
                    if '!Identifiers' in sbtab.columns[i]:
                        try:
                            db = re.search('Identifiers:(.*)',sbtab.columns[i])
                            url = 'http://identifiers.org/%s/%s' % (db.group(1), col)                        
                            html += '<td><a href="%s">%s</a></td>' % (url, col)
                        except: html += '<td>%s</td>' % (col)
                    else:
                        html += '<td>%s</td>' % (col)
            html += '</tr>'

        # comment rows
        for row in sbtab.comments:
            html += '<tr>'
            for col in row:
                html += '<td>%s</td>' % col
            html += '</tr>'

        # close table
        html += '</tbody></table>'

        return html
    
    ##############################################################################
    # read in header and footer from HTML template
    if mode == 'sbtab_online':
        p = os.path.join(os.path.dirname(__file__), '../modules/template_sbtab_online.html')
        try:
            html = open(p, 'r')
            html_template = html.read()
            html.close()
        except:
            print('HTML template was not found.')
            return False
    elif mode == 'standalone':
        html_template = False
        try_paths = ['html_templates/template_standalone.html',                     
                     os.path.join(os.path.dirname(__file__), '../html_templates/template_standalone.html'),
                     os.path.join(os.path.dirname(__file__), 'html_templates/template_standalone.html'),
                     template]
        for path in try_paths:
            try:
                html = open(path, 'r')
                html_template = html.read()
                html.close()
            except: pass
        if not html_template:
            print('HTML template was not found.')
            return False
    else:
        print('Invalid mode %s. Please use either "sbtab_online" or "standalone".' % mode)
        return False

    try:
        header = re.search('(<html lang="en">.*<main>)', html_template, re.DOTALL).group(0)
        footer = re.search('(</main>.*</html>)', html_template, re.DOTALL).group(0)
    except:
        print('Cannot read required template.html.')
        return False

    html = header

    try:
        ot = sbtab.object_type
    except:
        print('You have not provided a valid SBtab object as input.')
        return False
    
    # replace title placeholder with actual title
    if len(title_string):
        html = html.replace('TitlePlaceholder',title_string)
    else:
        html = html.replace('TitlePlaceholder',sbtab.filename)

    # read in definitions file for nice mouse over
    if mode == 'standalone' and len(definitions_file):
        sbtab_def = open_definitions_file(definitions_file)
    else:
        sbtab_def = open_definitions_file()
        
    # now build the html file
    if sbtab.object_type == 'table':
        html += _build_main(sbtab, sbtab_def)
    elif sbtab.object_type == 'doc':
        for sbtab in sbtab.sbtabs:
            html += _build_main(sbtab, sbtab_def) + '<br><hr>'
    else:
        print('The given SBtab object is invalid.')
        return False
            
    html += footer
    
    return html


def open_definitions_file(_path=None):
    '''
    Opens the SBtab definitions file, which can be in several locations.

    Parameters
    ----------
    _path: str
        Optional path to the definitions.tsv.

    Returns: SBtab.SBtabTable
        SBtab definitions file as SBtabTable object.        
    '''
    sbtab_def = False
    
    if _path: try_paths = [_path]
    else:
        try_paths = ['definitions.tsv',
                     os.path.join(os.path.dirname(__file__), '../static/files/default_files/definitions.tsv'),
                     os.path.join(os.path.dirname(__file__), '../definition_table/definitions.tsv'),
                     os.path.join(os.path.dirname(__file__), 'definitions.tsv'),
                     os.path.join(os.path.dirname(__file__), 'files/default_files/definitions.tsv')]

    for path in try_paths:
        try:
            def_file = open(path, 'r')
            file_content = def_file.read()
            sbtab_def = SBtab.SBtabTable(file_content, 'definitions.tsv')
            def_file.close()
            break
        except: pass

    return sbtab_def


def extract_supported_table_types():
    '''
    Extracts all allowed SBtab table types from the definitions file.

    Returns: list
        List of supported SBtab table types.
    '''
    sbtab_def = open_definitions_file()
    
    supported_types = []
    for row in sbtab_def.value_rows:
        t = row[sbtab_def.columns_dict['!Parent']]
        if t not in supported_types and t != 'SBtab':
            supported_types.append(t)

    return supported_types


def find_descriptions(def_file, table_type):
    '''
    Preprocesses the definitions file in order to enable some nice mouseover effects for the known column names.

    Parameters
    ----------
    def_file: SBtab.SBtabTable
        Definitions file as SBtab table object.
    table_type: str
        SBtab table type for which the descriptions shall be extracted.

    Returns: (dict, dict)
        Two dictionaries that link the column name to its description and to a Bool flag indicating if it
        is a shortname identifier possibly linking to another SBtab table.
    '''
    col2description = {}
    col2link = {}

    for row in def_file.value_rows:
        if row[def_file.columns_dict['!Parent']] == table_type:
            col2description[row[def_file.columns_dict['!Name']]] = row[def_file.columns_dict['!Description']]
            col2link['!'+row[def_file.columns_dict['!Name']]] = row[def_file.columns_dict['!isShortname']]

    return (col2description, col2link)


def xlsx_to_tsv(file_object, f='web'):
    '''
    Converts xlsx SBtab file to tsv format.

    Parameters
    ----------
    file_object: xlsx file object
        SBtab table as xlsx file object.
    f: str
        String indicating how the file object is represented ('web' for online file handling, otherwise normal file handling)
    
    Returns: str
        SBtab table as tsv string.
    '''
    import openpyxl
    from io import BytesIO

    if f == 'web': wb = openpyxl.load_workbook(filename = BytesIO(file_object))
    else: wb = openpyxl.load_workbook(filename = file_object)
    ws = wb.active
    ranges = wb[ws.title]
    table_string = ''

    for row in ranges:
        for column in row:
            if str(row[0].value).startswith('!'):
                if column.value != None and str(column.value) != '':
                    table_string += str(column.value) + '\t'
            else:
                table_string += str(column.value) + '\t'
        table_string = table_string[:-1] + '\n'
        
    return table_string


def tab_to_xlsx(sbtab_object):
    '''
    Converts SBtab object to xlsx file object.

    Parameters
    ----------
    sbtab_object: SBtab.SBtabTable
        SBtab table object.

    Returns: xlsx file object
        SBtab table as xlsx file object.
    '''
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws['A1'] = sbtab_object.header_row
    ws.append(sbtab_object.columns)
    for row in sbtab_object.value_rows:
        ws.append(row)

    wb.save('transition.xlsx')
    
    f = open('transition.xlsx','rb')
    fileobject = f.read()
    f.close()

    return fileobject

def id_checker(sbtab, sbml):
    '''
    this function checks, whether all the entries of the SBML ID columns of the SBtab file can also be
    found in the SBML file. If not, these are omitted during the balancing. But there should be a warning
    to raise user awareness.
    '''
    sbtabid2sbmlid = []

    reaction_ids_sbml = []
    species_ids_sbml  = []

    s_id = None
    r_id = None

    for reaction in sbml.getListOfReactions():
        reaction_ids_sbml.append(reaction.getId())
    for species in sbml.getListOfSpecies():
        species_ids_sbml.append(species.getId())

    for row in sbtab.value_rows:
        if len(row) < 3: continue
        try: s_id = sbtab.columns_dict['!Compound']
        except:
            try: s_id = sbtab.columns_dict['!Compound:SBML:species:id']
            except: sbtabid2sbmlid.append('Error: The SBtab file lacks the obligatory column "'"!Compound"'"/"'"!Compound:SBML:species:id"'" to link the parameter entries to the SBML model species.')

        try: r_id = sbtab.columns_dict['!Reaction']
        except:
            try: r_id = sbtab.columns_dict['!Reaction:SBML:reaction:id']
            except: sbtabid2sbmlid.append('Error: The SBtab file lacks the obligatory column "'"!Reaction"'"/"'"!Reaction:SBML:reaction:id"'" to link the parameter entries to the SBML model species.')
        try:
            if row[s_id] != '' and row[s_id] not in species_ids_sbml and row[s_id] != 'nan' and row[s_id] != 'None':
                sbtabid2sbmlid.append('Warning: The SBtab file holds a species ID which does not comply to any species ID in the SBML file: %s'%(row[s_id]))
        except: pass
        try:
            if row[r_id] != '' and row[r_id] not in reaction_ids_sbml and row[r_id] != 'nan' and row[r_id] != 'None':
                sbtabid2sbmlid.append('Warning: The SBtab file holds a reaction ID which does not comply to any reaction ID in the SBML file: %s'%(row[r_id]))
        except: pass
            
    return sbtabid2sbmlid


def xml_to_html(sbml_file):
    '''
    Generates HTML view of XML (SBML) file.

    Parameters
    ----------
    sbml_file: str
        SBML file in string representation.

    Returns: str
        SBML file as HTML view.
    '''
    old_sbml = sbml_file.split('\n')
    new_sbml = '<xmp>'
    for row in old_sbml:
        new_sbml += row + '\n'
    new_sbml += '</xmp>'

    return new_sbml


def tsv_to_html(sbtab, filename=None):
    '''
    generates html view out of tsv file
    '''
    sbtab_html = '''
    <html lang="en">
    <head>
    <meta charset="utf-8">
    <meta name="author" content="Timo Lubitz">
    <meta name="description"  content="Parameter Balancing Website">
    
    <!-- this is required for responsiveness -->
    <meta http-equiv="X-UA-Compatible" content="IE=edge">
    <meta name="viewport" content="width=device-width, initial-scale=1">

    <title>Parameter Balancing for Kinetic Models of Cell Metabolism</title>
    <!--<link rel="stylesheet" type="text/css" href="css/pb.css">-->
    <link href="../../static/css/css_template/css/bootstrap.min.css" rel="stylesheet">
    <link href="../../static/css/css_template/css/custom.css" rel="stylesheet">
    <link rel="shortcut icon" href="/pb/static/css/css_template/img/pb-logo.png" type="image/icon">
    <link rel="icon" href="/pb/static/css/css_template/img/pb-logo.png" type="image/icon">
    </head>

    <body>
    <!-- navbar: this is a navbar; navbar-inverse: it's dark; navbar-static-top: it's always at the top -->
    <nav class="navbar navbar-inverse navbar-fixed-top">
    <!-- setting a max-width by using a container -->
      <div class="container">
        <div class="navbar-header">
	  <!-- button is hidden on desktop, becomes a hamburger on mobile! the span items are the hamburger lines --> 
	  <button type="button" class="navbar-toggle collapsed" data-toggle="collapse" data-target="#bs-example-navbar-collapse-1" aria-expanded="false">
	    <span class="sr-only">Toggle navigation</span>
	    <span class="icon-bar"></span>
	    <span class="icon-bar"></span>
	    <span class="icon-bar"></span>
	  </button>
	  <a class="navbar-brand" href="#">Parameter Balancing</a>
	</div>
	
	<!-- simple right aligned list-->
	<div class="collapse navbar-collapse" id="bs-example-navbar-collapse-1">
	  <ul class="nav navbar-nav navbar-right">
            <li> <a href="../../static/css/css_template/gettingstarted.html" title="Getting started">Getting Started</a></li>
	    <li> <a href="../../default/balancing.html" title="Go to online balancing">Online Balancing</a></li>
	    <li> <a href="../../static/css/css_template/documentation.html" title="Documentation and Manuals">Documentation</a></li>
	    <li> <a href="../../static/css/css_template/download.html" title="Installation and Downloads">Download</a></li>
	    <li> <a href="../../static/css/css_template/contact.html" title="Contact the balancing team">Contact</a></li>
	  </ul>
	</div>
      </div>
    </nav>

    <header>
    '''
    if type(sbtab) == str and filename:
        ugly_sbtab = sbtab.split('\n')
        #nice_sbtab = '<p><h2><b>%s</b></h2></p>' % filename
        sbtab_html += '<h2 style="padding-top:50px" align="center"><small>%s</small></h2></div></header>' % filename
        delimiter = check_delimiter(sbtab)
    else:
        ugly_sbtab = sbtab.return_table_string().split('\n')
        #nice_sbtab = '<p><h2><b>'+sbtab.filename+'</b></h2></p>'
        sbtab_html += '<h2 style="padding-top:50px" align="center"><small>'+sbtab.filename+'</small></h2></div></header>'
        delimiter = sbtab.delimiter

    sbtab_html += '''
    <main>
    <div class="container-fluid bg-1c text-center">
    <div class="row" style="background-color:#9d9d9d;">
    <div class="col-sm-1"></div>
    <div class="col-sm-10">
    <table class="table-striped" style="font-size:13px;background-color:#fff;padding:3px;">'''
        
    first = True
    for i, row in enumerate(ugly_sbtab):
        # declaration of first SBtab in document
        if row.startswith('!!') and first:
            sbtab_html += '<tr><th colspan="%s" style="padding:3px;">%s</th></tr>' % (len(ugly_sbtab[i+2]), row)
            first = False

        # conclusion of SBtab and beginning of new SBtab (if there are more than one)
        elif row.startswith('!!'):
            sbtab_html += '</table><br><table class="table-striped" style="font-size:13px;background-color:#fff;">'
            sbtab_html += '<tr><th colspan="%s" style="padding:3px;">%s</th></tr>' % (len(ugly_sbtab[i+2]), row)

        # column header row
        elif row.startswith('!'):
            splitrow = row.split(delimiter)
            sbtab_html += '<tr>'
            for col in splitrow:
                sbtab_html += '<th style="padding:3px;">%s</th>' % col
            sbtab_html += '</tr>'

        # comment row
        elif row.startswith('%'):
            sbtab_html += '<tr bgcolor="#C0C0C0">%s</tr>' % row

        # log file header
        elif row.startswith('Parameter balancing log file'):
            sbtab_html += '<tr>%s</tr>'

        # normal row
        else:
            splitrow = row.split(delimiter)
            sbtab_html += '<tr>'
            for col in splitrow:
                sbtab_html += '<td style="padding:3px;">%s</td>' % col
            sbtab_html += '</tr>'

        '''
        # normal row
        for i,thing in enumerate(row.split(delimiter)):
            if thing.startswith('!!'): continue
            #new_row = '<td>'+str(thing)+'</td>'
            #nice_sbtab += new_row
        #nice_sbtab += '</tr>'
        '''
    sbtab_html += '''
    </table>
    </div>
    <div class="col-sm-1"></div>
    </div>
    </div>

    </main>
    <hr>
    <footer class="container-fluid bg-3 text-center">
    <p>Thanks to <a href="https://getbootstrap.com/" target="_blank">Bootstrap</a> and <a href="http://web2py.com/" target="_blank">Web2py</a>. Code and further information on <a href="https://github.com/tlubitz/parameter_balancing" target="_blank">Github</a>.</p> 
    </footer>
    
    <script src="https://ajax.googleapis.com/ajax/libs/jquery/1.11.3/jquery.min.js"></script>
    <script src="js/bootstrap.min.js"></script>
    </body>
    </html>
    '''
    return sbtab_html
